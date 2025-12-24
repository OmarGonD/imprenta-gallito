import openpyxl
import shutil

excel_path = r'D:\web_proyects\imprenta_gallito\static\data\products_complete.xlsx'
backup_path = r'D:\web_proyects\imprenta_gallito\static\data\products_complete_backup_snacks.xlsx'

# Marketing templates based on packaging types and candy styles seen on the website
marketing_templates = [
    {
        "name": "Lata Metálica Premium",
        "description": "Elegante lata metálica personalizable con dulces selectos. Perfecta para regalos corporativos de alto impacto. Tu logo grabado en un empaque que transmite calidad y distinción."
    },
    {
        "name": "Bolsa Gourmet Transparente",
        "description": "Bolsa de dulces gourmet con presentación cristalina. Ideal para eventos y ferias. La transparencia permite mostrar la calidad del producto mientras destaca tu marca con elegancia."
    },
    {
        "name": "Frasco de Vidrio Ejecutivo",
        "description": "Frasco de vidrio con tapa metálica, lleno de dulces premium. Regalo corporativo sofisticado y reutilizable. Un detalle que perdura en escritorios ejecutivos con tu marca siempre visible."
    },
    {
        "name": "Lata Corazón Especial",
        "description": "Lata en forma de corazón con dulces variados. Ideal para campañas de San Valentín, aniversarios corporativos y eventos especiales. Tu marca asociada con momentos memorables."
    },
    {
        "name": "Tubo Transparente Moderno",
        "description": "Tubo cilíndrico transparente con dulces coloridos. Diseño moderno y funcional que maximiza la visibilidad de tu marca. Perfecto para promociones masivas con estilo contemporáneo."
    },
    {
        "name": "Sobre Pillow Pack Clásico",
        "description": "Empaque tipo almohada con dulces individuales. Solución económica y práctica para eventos masivos. Tu logo impreso en un formato compacto que viaja con tus clientes."
    },
    {
        "name": "Lata Rectangular Profesional",
        "description": "Lata rectangular con bisagra y mentas de alta calidad. Presencia ejecutiva en bolsillos y escritorios. Tu marca acompañando a profesionales en su día a día."
    },
    {
        "name": "Mix de Frutos Secos Energy",
        "description": "Mezcla nutritiva de nueces, almendras y semillas en empaque personalizado. Ideal para empresas wellness y eventos deportivos. Salud y energía con tu marca como protagonista."
    },
    {
        "name": "Gomas Coloridas Diversión",
        "description": "Ositos de goma y dulces coloridos en presentación vibrante. Perfectos para campañas familiares y eventos infantiles. Alegría y diversión que asocian tu marca con momentos felices."
    },
    {
        "name": "Chocolates Premium Selection",
        "description": "Selección de chocolates finos en empaque elegante. Regalo de lujo para clientes VIP y ocasiones especiales. Tu marca elevada a la categoría premium del merchandising."
    },
    {
        "name": "Mentas Refrescantes Ejecutivas",
        "description": "Mentas de alta calidad en empaque discreto y profesional. El acompañante perfecto en reuniones de negocios. Frescura y profesionalismo que refuerzan tu imagen corporativa."
    },
    {
        "name": "Lata Redonda Vintage",
        "description": "Lata circular con diseño clásico, llena de dulces tradicionales. Nostalgia y calidad en un solo empaque. Tu marca con un toque retro que conecta emocionalmente."
    },
    {
        "name": "Bolsa con Lazo Regalo",
        "description": "Bolsa decorativa con lazo de seda y dulces gourmet. Presentación de regalo lista para entregar. Máximo impacto visual con mínimo esfuerzo, tu marca como el mejor detalle."
    },
    {
        "name": "Caja Snack Individual",
        "description": "Pequeña caja de cartón con snack personalizado. Ideal para welcome packs y eventos corporativos. Practicidad y branding en un formato compacto y accesible."
    },
    {
        "name": "Caramelos Artesanales Gourmet",
        "description": "Caramelos de fabricación artesanal en empaque premium. Sabores únicos que destacan la exclusividad de tu marca. Perfecto para clientes que valoran la diferenciación."
    },
    {
        "name": "Set Dulces Variados Fiesta",
        "description": "Surtido de dulces y caramelos en presentación festiva. Ideal para celebraciones corporativas y fin de año. Variedad y color que transmiten alegría con tu logo como anfitrión."
    },
    {
        "name": "Contenedor Plástico Eco-Friendly",
        "description": "Recipiente reutilizable con dulces naturales. Compromiso ambiental y sabor en perfecta armonía. Tu marca demostrando responsabilidad social de manera deliciosa."
    },
    {
        "name": "Galletas Gourmet Artesanales",
        "description": "Galletas de receta especial en lata decorativa. Exclusividad y sabor casero para obsequios memorables. Tu marca asociada con la calidez de lo hecho a mano."
    },
    {
        "name": "Lata Ovalada Sofisticada",
        "description": "Empaque ovalado con acabado mate y dulces selectos. Diseño vanguardista que rompe con lo tradicional. Tu marca innovadora hasta en el detalle más pequeño."
    },
    {
        "name": "FlowWrap Snack Práctico",
        "description": "Empaque flow-wrap individual para máxima higiene. Ideal para distribución masiva en eventos y conferencias. Tu marca en cada mano de manera práctica y profesional."
    }
]

try:
    # Create backup
    shutil.copy2(excel_path, backup_path)
    print(f"✓ Backup created: {backup_path}\n")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Find column indices
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value:
            headers[cell_value] = col
    
    # Get column indices
    name_col = headers.get('product_name')
    desc_col = headers.get('description')
    subcat_col = headers.get('subcategory_slug')
    sku_col = headers.get('sku')
    
    if not all([name_col, desc_col, subcat_col, sku_col]):
        print("ERROR: Missing required columns")
        exit(1)
    
    # Update rows
    updated_count = 0
    print("=" * 80)
    print("UPDATING SNACKS & CANDIES PRODUCTS")
    print("=" * 80 + "\n")
    
    for row in range(2, ws.max_row + 1):
        subcategory = ws.cell(row, subcat_col).value
        
        if subcategory and str(subcategory).strip() == 'snacks-caramelos':
            # Get SKU to extract code
            sku = str(ws.cell(row, sku_col).value or '')
            code_parts = sku.split('-')
            if len(code_parts) >= 3:
                # Format: PP-SNK-0001 -> get "0001"
                code = code_parts[-1]
            else:
                code = f"{updated_count+1:04d}"
            
            # Get template (rotate through the 20 templates)
            template = marketing_templates[updated_count % len(marketing_templates)]
            
            # Update cells
            new_name = f"{template['name']} {code}"
            ws.cell(row, name_col).value = new_name
            ws.cell(row, desc_col).value = template['description']
            
            updated_count += 1
            print(f"✓ {updated_count:2d}. {new_name}")
    
    # Save
    wb.save(excel_path)
    wb.close()
    
    print("\n" + "=" * 80)
    print(f"✅ Successfully updated {updated_count} snacks & candies products!")
    print("=" * 80)
    
except PermissionError as e:
    print(f"\n❌ Permission Error: The file is likely open in Excel.")
    print(f"Please close the file and try running this script again.")
except Exception as e:
    print(f"\n❌ Error: {e}")
    import traceback
    traceback.print_exc()
