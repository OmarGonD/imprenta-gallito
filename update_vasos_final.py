import openpyxl
import shutil
from datetime import datetime

excel_path = r'D:\web_proyects\imprenta_gallito\static\data\products_complete.xlsx'
backup_path = r'D:\web_proyects\imprenta_gallito\static\data\products_complete_backup.xlsx'

# Marketing templates based on the glass styles seen on the website
marketing_templates = [
    {
        "name": "Vaso Pilsner Elegante",
        "description": "Vaso cervecero estilo pilsner de cristal transparente. Ideal para bares, restaurantes y eventos corporativos. Su forma estilizada realza cualquier bebida y tu logo destacará con elegancia."
    },
    {
        "name": "Copa Sin Tallo Premium",
        "description": "Copas modernas sin tallo, perfectas para vino o cócteles. Diseño contemporáneo que combina funcionalidad y estilo. Excelente para regalos corporativos y merchandising de alta gama."
    },
    {
        "name": "Vaso Tumbler Versátil",
        "description": "Vaso tipo tumbler de vidrio transparente con forma cónica. Perfecto para cualquier tipo de bebida. Su diseño clásico garantiza que tu marca sea visible desde cualquier ángulo."
    },
    {
        "name": "Copa Cóctel Profesional",
        "description": "Copa para cócteles con base robusta y diseño profesional. Ideal para hoteles, eventos y promociones de bebidas premium. Dale a tu marca una imagen sofisticada y memorable."
    },
    {
        "name": "Jarra Cervecera con Asa",
        "description": "Jarra de cerveza tipo stein con asa ergonómica. Resistente y de gran capacidad, perfecta para festivales, oktoberfest y eventos temáticos. Tu logo será el protagonista en cada brindis."
    },
    {
        "name": "Vaso Alto Cristalino",
        "description": "Vaso alto de cristal transparente ideal para refrescos, cervezas y bebidas mixtas. Su elegante altura maximiza el espacio para personalización. Perfecto para campañas de gran alcance."
    },
    {
        "name": "Copa Vino Stemless Duo",
        "description": "Set de copas de vino sin tallo, tendencia moderna en vajilla. Perfecto para regalos ejecutivos y kits de bienvenida. Su diseño minimalista realza la presencia de tu marca."
    },
    {
        "name": "Vaso Cónico Base Gruesa",
        "description": "Vaso cónico con base reforzada que transmite calidad y durabilidad. Excelente estabilidad y peso ideal. Tu imagen corporativa lucirá profesional en cada mesa o escritorio."
    },
    {
        "name": "Vaso Everyday Promocional",
        "description": "Vaso de uso diario con diseño práctico y atemporal. Ideal para campañas masivas y regalos promocionales accesibles. Mantén tu marca presente en la rutina de tus clientes."
    },
    {
        "name": "Copa Elegance Premium",
        "description": "Copa de vidrio con diseño elegante y refinado. Perfecta para eventos especiales, bodas corporativas y premios. Asocia tu marca con momentos de celebración y exclusividad."
    }
]

try:
    # Create backup
    shutil.copy2(excel_path, backup_path)
    print(f"✓ Backup created: {backup_path}")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Find column indices
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value:
            headers[cell_value] = col
    
    print(f"Found columns: {list(headers.keys())}")
    
    # Get column indices
    name_col = headers.get('product_name')
    desc_col = headers.get('description')
    subcat_col = headers.get('subcategory_slug')
    sku_col = headers.get('sku')
    
    if not all([name_col, desc_col, subcat_col, sku_col]):
        print("ERROR: Missing required columns")
        print(f"name_col={name_col}, desc_col={desc_col}, subcat_col={subcat_col}, sku_col={sku_col}")
        exit(1)
    
    # Update rows
    updated_count = 0
    for row in range(2, ws.max_row + 1):
        subcategory = ws.cell(row, subcat_col).value
        
        if subcategory and str(subcategory).strip() == 'vasos':
            # Get SKU to extract code
            sku = str(ws.cell(row, sku_col).value or '')
            code_parts = sku.split('-')
            if len(code_parts) >= 3:
                code = code_parts[-1]
            else:
                code = f"#{updated_count+1:02d}"
            
            # Get template
            template = marketing_templates[updated_count % len(marketing_templates)]
            
            # Update cells
            new_name = f"{template['name']} {code}"
            ws.cell(row, name_col).value = new_name
            ws.cell(row, desc_col).value = template['description']
            
            updated_count += 1
            print(f"✓ Row {row}: {new_name}")
    
    # Save
    wb.save(excel_path)
    wb.close()
    
    print(f"\n✅ Successfully updated {updated_count} products!")
    print(f"Backup saved to: {backup_path}")
    
except PermissionError as e:
    print(f"❌ Permission Error: The file is likely open in Excel or another program.")
    print(f"Please close the file and try running this script again.")
    print(f"Error: {e}")
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
